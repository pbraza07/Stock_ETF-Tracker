from unittest.mock import patch
from pathlib import Path
from io import BytesIO
import numpy as np
import pandas as pd
import pytest
from pandas.testing import assert_frame_equal
from top12_rankings import (
    build_top12_rankings,
    select_top12,
    path_metrics,
    walk_forward_rankings,
)
from top12_history import record_run, merge_ledgers
from top12_ui import portfolio_inputs
from top12_exports import build_top12_excel, build_top12_pdf

YEARS = [str(y) for y in range(2006, 2026)]


def fixture():
    rng = np.random.default_rng(77)
    rows = []
    for i in range(36):
        row = {
            "Symbol": f"T{i:02}",
            "Name": f"Test Company {i}",
            "Sector": f"Sector {i//6}",
            "Type": "Stock",
            "Price": 100.0,
            "6M": float(i),
        }
        row.update({y: float(rng.normal(9, 20)) for y in YEARS})
        rows.append(row)
    rows.append(dict(rows[0], Symbol="ETF", Type="ETF"))
    rows.append(dict(rows[0], Symbol="INVALID", Sector="Unknown"))
    return pd.DataFrame(rows)


@pytest.fixture(scope="module")
def result():
    return build_top12_rankings(fixture(), YEARS, simulations=100)


def test_all_eligible_and_caps(result):
    assert len(result["all_scores"]) == 36
    for kind in ("Recession", "Max Profit"):
        table = result[kind]
        assert len(table) == 12
        assert table.groupby("Sector").size().max() <= 4
        assert table.Symbol.nunique() == 12
        assert table[kind + " Score"].between(0, 100).all()


def test_seed_and_percentiles(result):
    again = build_top12_rankings(fixture(), YEARS, simulations=100)
    assert_frame_equal(result["Recession"], again["Recession"])
    for prefix in ("", "Bear "):
        values = result["all_scores"][
            [prefix + f"P{q} Future Return %" for q in (10, 25, 50, 75, 90)]
        ].to_numpy()
        assert (np.diff(values, axis=1) >= 0).all()


def test_hysteresis_and_insufficient_diversity():
    frame = pd.DataFrame(
        {
            "Symbol": [str(i) for i in range(13)],
            "Sector": [str(i % 4) for i in range(13)],
            "Score": [100.0 - i for i in range(13)],
        }
    )
    initial = select_top12(frame, "Score")
    frame.loc[12, "Score"] = 89.5
    assert "12" not in set(select_top12(frame, "Score", initial, 1).Symbol)
    frame.loc[12, "Score"] = 91
    assert "12" in set(select_top12(frame, "Score", initial, 1).Symbol)
    with pytest.raises(ValueError):
        select_top12(frame.assign(Sector="One"), "Score")


def test_append_only_history(result):
    table = result["Recession"]
    ledger = record_run({}, "Recession", table, {}, "2026-09-01T00:00:00Z")
    assert record_run(ledger, "Recession", table, {}, "2026-09-02T00:00:00Z") == ledger
    changed = table.copy()
    changed.loc[0, "Symbol"] = "NEW"
    updated = record_run(ledger, "Recession", changed, {}, "2026-09-02T00:00:00Z")
    assert updated["events"][:12] == ledger["events"]
    assert merge_ledgers(updated, ledger) == updated


def test_portfolio_and_exports(result):
    from openpyxl import load_workbook

    for kind in ("Recession", "Max Profit"):
        for mode in ("Equal Weight", "Score Weighted"):
            inputs = portfolio_inputs(result[kind], kind, mode, 300000, 10)
            assert (
                len(inputs["holdings"]) == 12
                and abs(sum(inputs["allocations"].values()) - 100) < 1e-9
            )
            assert inputs["strategy"] == "Both"
    xlsx = build_top12_excel("Recession", result["Recession"], result)
    assert "All Candidate Scores" in load_workbook(BytesIO(xlsx)).sheetnames
    assert build_top12_pdf("Recession", result["Recession"], result).startswith(b"%PDF")


def test_walkforward_truncates_evidence():
    import top12_rankings as rankings

    original = rankings.score_universe
    seen = []

    def checking(market, years, *args, **kwargs):
        assert not {"Price", "6M", "YTD", "MarketCap"} & set(market)
        assert max(int(c) for c in market if c.isdigit()) == max(map(int, years))
        seen.append(max(map(int, years)))
        return original(market, years, *args, **kwargs)

    with patch.object(rankings, "score_universe", checking):
        data = walk_forward_rankings(fixture(), YEARS)
    assert seen and not data.empty
    assert (data["As Of"] < data["Evaluation Start"]).all()


def test_path_depletion_recovery():
    dd, recovery = path_metrics([-0.5, 1.0])
    assert dd == -50 and recovery == 2


def test_export_cache_invalidation():
    import runtime_performance as r

    r.projection_exports.clear()
    with patch.object(
        r, "build_excel_export", return_value=b"x"
    ) as count, patch.object(r, "build_csv_export", return_value=b"c"), patch.object(
        r, "build_pdf_export", return_value=b"p"
    ):
        r.projection_exports({"v": 1})
        r.projection_exports({"v": 1})
        assert count.call_count == 1
        r.projection_exports({"v": 2})
        assert count.call_count == 2
    r.projection_exports.clear()


def test_monthly_stress_uses_exact_window():
    from top12_rankings import stress_evidence

    hist = pd.Series(
        0.01, index=[str(p) for p in pd.period_range("2018-01", "2023-12", freq="M")]
    )
    hist.loc[["2020-02", "2020-03"]] = -0.2
    annual = pd.Series(0.2, index=YEARS)
    info = stress_evidence(hist, annual, hist, [hist])
    assert info["Stress Evidence Basis"] == "Monthly observed"
    assert info["Worst Stress Return %"] == pytest.approx(-36)
    assert info["Stress Market Excess %"] == pytest.approx(0)
    assert info["Stress Sector Excess %"] == pytest.approx(0)
    assert np.isfinite(info["Post Trough 6M Return %"])


def test_twelve_stock_projection_has_both_sequential_paths(result):
    from future_projection import run_future_projection

    inputs = portfolio_inputs(
        result["Max Profit"], "Max Profit", "Score Weighted", 300000, 3
    )
    inputs["simulation_count"] = 100
    projected = run_future_projection(fixture(), inputs, YEARS, live_context={})
    assert set(projected["strategies"]) == {"Rebalanced", "Non-Rebalanced"}
    for payload in projected["strategies"].values():
        assert len(payload["table"]) == 3
        s = payload["summary"]
        assert all(
            s[f"P{a} Ending Balance"] <= s[f"P{b} Ending Balance"]
            for a, b in zip((10, 25, 50, 75), (25, 50, 75, 90))
        )
        assert 0 <= s["Probability Positive Total Wealth"] <= 100


def test_sector_aliases_cannot_evade_cap():
    from top12_rankings import score_universe

    data = fixture()
    data.loc[:5, "Sector"] = "Financials"
    data.loc[6:11, "Sector"] = "Financial Services"
    frame, _, _ = score_universe(data, YEARS, simulations=100)
    assert "Financials" not in set(frame.Sector)
    picks = select_top12(frame, "Recession Score")
    assert picks.Sector.eq("Finance").sum() <= 4


def test_lazy_tabs_keep_inputs():
    from streamlit.testing.v1 import AppTest

    source = """
import streamlit as st
from runtime_performance import preserve_navigation_state
preserve_navigation_state()
a,b=st.tabs(['Main','Future Projection'],key='workspace_navigation',on_change='rerun')
if b.open:
    with b: st.text_input('Investment',key='fp_test')
"""
    app = AppTest.from_string(source).run()
    assert len(app.text_input) == 0
    app.session_state["workspace_navigation"] = "Future Projection"
    app.run()
    app.text_input(key="fp_test").set_value("123456").run()
    app.session_state["workspace_navigation"] = "Main"
    app.run()
    app.session_state["workspace_navigation"] = "Future Projection"
    app.run()
    assert app.text_input(key="fp_test").value == "123456"
    assert not app.exception
