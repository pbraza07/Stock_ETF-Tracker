"""Future Projection calculations and exports for MarketScope.

This module is intentionally independent from Streamlit.  The existing
historical simulators remain untouched; the UI supplies the current MarketScope
universe plus actual monthly returns when available, and this module performs a
forward-looking, correlated, regime-aware Monte Carlo projection.
"""

from __future__ import annotations

import hashlib
import io
import json
import math
from dataclasses import dataclass
from datetime import date, datetime
from typing import Callable, Iterable

import numpy as np
import pandas as pd

from future_projection_config import (
    MODEL_LIMITATIONS,
    capital_market_assumptions,
    model_defaults,
)


ProgressCallback = Callable[[int, int, str], None]


class ProjectionValidationError(ValueError):
    """A user-correctable projection validation failure."""


@dataclass
class PreparedProjectionModel:
    symbols: list[str]
    names: list[str]
    types: list[str]
    categories: list[str]
    expected_annual_returns: np.ndarray
    annual_covariance: np.ndarray
    period_log_returns: np.ndarray
    period_covariance: np.ndarray
    period_frequency: str
    confidence: str
    credible: bool
    warnings: list[str]
    holding_assumptions: pd.DataFrame
    historical_returns: pd.DataFrame
    diagnostics: dict


def parse_currency(value) -> float:
    """Parse currency-like input while accepting commas, spaces, and `$`."""

    if value is None:
        return 0.0
    if isinstance(value, (int, float, np.number)):
        parsed = float(value)
    else:
        text = str(value).strip().replace("$", "").replace(",", "").replace(" ", "")
        if text in {"", ".", "-", "+"}:
            return 0.0
        try:
            parsed = float(text)
        except (TypeError, ValueError) as exc:
            raise ProjectionValidationError(f"'{value}' is not a valid currency amount.") from exc
    if not np.isfinite(parsed):
        raise ProjectionValidationError("Currency amounts must be finite numbers.")
    return parsed


def _number(value, default: float = 0.0) -> float:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return float(default)
    return parsed if np.isfinite(parsed) else float(default)


def _normalized_market(market: pd.DataFrame) -> pd.DataFrame:
    if market is None or market.empty or "Symbol" not in market.columns:
        return pd.DataFrame()
    out = market.copy()
    out["Symbol"] = out["Symbol"].astype(str).str.strip().str.upper()
    out = out[out["Symbol"].ne("")].drop_duplicates("Symbol", keep="last")
    return out.set_index("Symbol", drop=False)


def holding_category(row: pd.Series) -> str:
    """Use stock sector, but preserve an ETF asset class/fund category."""

    instrument_type = str(row.get("Type") or "Stock").strip().upper()
    if instrument_type == "ETF":
        for key in ("ETF Category", "Fund Category", "Asset Class", "Category", "Sector", "Industry"):
            value = str(row.get(key) or "").strip()
            if value and value.lower() not in {"nan", "unknown", "etf / fund"}:
                return value
        return "ETF / Fund"
    value = str(row.get("Sector") or "Unknown").strip()
    return value if value and value.lower() != "nan" else "Unknown"


def normalize_projection_inputs(inputs: dict) -> dict:
    """Normalize UI/persisted values into the engine's stable public contract."""

    raw = dict(inputs or {})
    holdings = [str(x or "").strip().upper() for x in raw.get("holdings", [])]
    allocations_raw = raw.get("allocations") or {}
    if isinstance(allocations_raw, (list, tuple)):
        allocations = {
            symbol: _number(allocations_raw[idx], 0.0)
            for idx, symbol in enumerate(holdings)
            if idx < len(allocations_raw)
        }
    else:
        allocations = {str(k).upper(): _number(v, 0.0) for k, v in dict(allocations_raw).items()}

    frequency = str(raw.get("withdrawal_frequency") or "Yearly").title()
    if frequency not in {"No Withdrawal", "Yearly", "Monthly"}:
        frequency = "Yearly"
    strategy = str(raw.get("strategy") or "Both").title().replace("Non-Rebalanced", "Non-Rebalanced")
    if strategy not in {"Rebalanced", "Non-Rebalanced", "Both"}:
        strategy = "Both"
    quality = str(raw.get("scenario_quality") or "Advanced")
    defaults = model_defaults()
    simulation_count = int(raw.get("simulation_count") or defaults["simulation_counts"].get(quality, 20_000))
    normalized = {
        "starting_investment": parse_currency(raw.get("starting_investment", 400_000.0)),
        "withdrawal_frequency": frequency,
        "annual_withdrawal": parse_currency(raw.get("annual_withdrawal", 200_000.0)),
        "monthly_withdrawal": parse_currency(raw.get("monthly_withdrawal", 200_000.0 / 12.0)),
        "withdrawal_timing": str(raw.get("withdrawal_timing") or "End of period"),
        "future_years": int(raw.get("future_years") or 10),
        "holdings": holdings,
        "allocation_mode": str(raw.get("allocation_mode") or "Equal Split"),
        "allocations": allocations,
        "strategy": strategy,
        "rebalancing_frequency": str(raw.get("rebalancing_frequency") or "Yearly").title(),
        "scenario_quality": quality,
        "simulation_count": simulation_count,
        "inflation_adjust_withdrawals": bool(raw.get("inflation_adjust_withdrawals", False)),
        "withdrawal_inflation_rate": _number(raw.get("withdrawal_inflation_rate"), 0.025),
        "annual_management_fee": _number(raw.get("annual_management_fee"), 0.0),
        "additional_contribution": parse_currency(raw.get("additional_contribution", 0.0)),
        "random_seed": int(raw.get("random_seed") if raw.get("random_seed") is not None else 20260903),
        "include_no_withdrawal_comparison": bool(raw.get("include_no_withdrawal_comparison", True)),
        "show_extended_range": bool(raw.get("show_extended_range", False)),
        "capital_market_assumption_date": str(
            raw.get("capital_market_assumption_date")
            or capital_market_assumptions()["broad_market_annual_geometric_return"]["as_of_date"]
        ),
        "forecast_start_year": int(raw.get("forecast_start_year") or datetime.utcnow().year),
    }
    if normalized["allocation_mode"].lower().startswith("equal") and len(holdings) == 4:
        normalized["allocations"] = {symbol: 25.0 for symbol in holdings}
    return normalized


def projection_payload_from_simulator(state: dict, symbols: list[str] | None = None) -> dict:
    """Translate the existing historical simulator state without changing it."""

    source = dict(state or {})
    holdings = [str(x).upper() for x in (symbols if symbols is not None else source.get("portfolio_symbols") or [])]
    holdings = list(dict.fromkeys(symbol for symbol in holdings if symbol))[:4]
    yearly_enabled = bool(source.get("portfolio_withdrawals_enabled"))
    monthly_enabled = bool(source.get("portfolio_monthly_withdrawals_enabled"))
    if monthly_enabled:
        frequency = "Monthly"
    elif yearly_enabled:
        frequency = "Yearly"
    else:
        frequency = "No Withdrawal"
    allocation_mode = "Custom Allocation" if str(source.get("portfolio_allocation_mode") or "").lower().startswith("custom") else "Equal Split"
    if allocation_mode == "Equal Split" and len(holdings) == 4:
        allocations = {symbol: 25.0 for symbol in holdings}
    else:
        allocations = {
            symbol: _number(source.get(f"portfolio_weight_{symbol}"), 0.0)
            for symbol in holdings
        }
    start = _number(source.get("portfolio_total_amount"), 400_000.0)
    if start <= 0:
        start = 400_000.0
    annual = _number(source.get("portfolio_annual_withdrawal"), 200_000.0)
    monthly = _number(source.get("portfolio_monthly_withdrawal"), 200_000.0 / 12.0)
    return {
        "starting_investment": start,
        "withdrawal_frequency": frequency,
        "annual_withdrawal": max(0.0, annual),
        "monthly_withdrawal": max(0.0, monthly),
        "withdrawal_timing": "End of period",
        "future_years": 10,
        "holdings": holdings,
        "allocation_mode": allocation_mode,
        "allocations": allocations,
        "strategy": "Both",
        "rebalancing_frequency": "Yearly",
        "scenario_quality": "Advanced",
    }


def validate_projection_inputs(inputs: dict, market: pd.DataFrame | None = None) -> tuple[list[str], list[str]]:
    """Return user-facing errors and warnings without exposing stack traces."""

    try:
        value = normalize_projection_inputs(inputs)
    except (ProjectionValidationError, TypeError, ValueError) as exc:
        return [str(exc)], []
    errors: list[str] = []
    warnings: list[str] = []
    if value["starting_investment"] <= 0:
        errors.append("Starting Investment must be greater than zero.")
    elif value["starting_investment"] < 1_000:
        errors.append("Starting Investment must be at least $1,000.")
    if value["annual_withdrawal"] < 0 or value["monthly_withdrawal"] < 0:
        errors.append("Withdrawal amounts cannot be negative.")
    if value["additional_contribution"] < 0:
        errors.append("Additional contributions cannot be negative.")
    if not 1 <= value["future_years"] <= 50:
        errors.append("Future Years must be between 1 and 50.")
    if value["withdrawal_timing"] not in {"End of period", "Beginning of period"}:
        errors.append("Withdrawal Timing must be Beginning of period or End of period.")
    if value["rebalancing_frequency"] not in {"Yearly", "Quarterly", "Monthly"}:
        errors.append("Rebalancing Frequency must be Yearly, Quarterly, or Monthly.")
    holdings = value["holdings"]
    if len(holdings) != 4 or any(not symbol for symbol in holdings):
        errors.append("Select all four Stock/ETF holdings before running the projection.")
    selected_holdings = [symbol for symbol in holdings if symbol]
    if len(set(selected_holdings)) != len(selected_holdings):
        errors.append("Each holding must use a different ticker; remove the duplicate ticker.")
    if len(holdings) == 4:
        allocation_total = sum(_number(value["allocations"].get(symbol), 0.0) for symbol in holdings)
        if abs(allocation_total - 100.0) > 1e-8:
            errors.append(f"Custom allocations must equal exactly 100%; the current total is {allocation_total:.2f}%.")
        if any(_number(value["allocations"].get(symbol), 0.0) < 0 for symbol in holdings):
            errors.append("Allocation percentages cannot be negative.")
    if not 1 <= value["simulation_count"] <= 100_000:
        errors.append("Simulation count must be between 1 and 100,000.")
    if value["annual_management_fee"] < 0 or value["annual_management_fee"] >= 1:
        errors.append("Annual management fee must be at least 0% and below 100%.")
    if value["withdrawal_inflation_rate"] <= -1:
        errors.append("Withdrawal inflation must be greater than -100%.")

    lookup = _normalized_market(market) if market is not None else pd.DataFrame()
    if market is not None:
        if lookup.empty:
            errors.append("The MarketScope stock/ETF dataset could not be loaded.")
        else:
            missing = [symbol for symbol in holdings if symbol and symbol not in lookup.index]
            if missing:
                errors.append("These holdings are not in the currently loaded MarketScope universe: " + ", ".join(missing))
            # The form always has four slots, but unselected slots are empty strings.
            # Do not index the market universe until all four real, unique tickers
            # have been selected and confirmed present in the loaded dataset.
            if (
                len(holdings) == 4
                and all(holdings)
                and len(set(holdings)) == 4
                and not missing
            ):
                completed_year_columns = [column for column in lookup.columns if str(column).isdigit()]
                observed_periods = 0
                for symbol in holdings:
                    row = lookup.loc[symbol]
                    if isinstance(row, pd.DataFrame):
                        row = row.iloc[-1]
                    values = pd.to_numeric(
                        pd.Series([row.get(column) for column in completed_year_columns]),
                        errors="coerce",
                    )
                    observed_periods += int((values.notna() & np.isfinite(values) & (values > -100.0)).sum())
                minimum = int(model_defaults()["minimum_credible_annual_periods"])
                if len(completed_year_columns) < minimum or observed_periods < minimum:
                    errors.append(
                        "The selected holdings have too little completed historical data for a credible projection. "
                        f"Choose holdings with at least {minimum} observed completed annual periods in total."
                    )
            stock_sectors: dict[str, list[str]] = {}
            for symbol in holdings:
                if symbol not in lookup.index:
                    continue
                row = lookup.loc[symbol]
                if isinstance(row, pd.DataFrame):
                    row = row.iloc[-1]
                if str(row.get("Type") or "Stock").upper() != "ETF":
                    stock_sectors.setdefault(holding_category(row), []).append(symbol)
            repeated = {sector: syms for sector, syms in stock_sectors.items() if sector != "Unknown" and len(syms) > 1}
            if repeated:
                detail = "; ".join(f"{sector}: {', '.join(syms)}" for sector, syms in repeated.items())
                warnings.append("Diversification warning - multiple stocks share a sector (" + detail + ").")
    if value["scenario_quality"] == "High Precision":
        warnings.append("High Precision runs 50,000 simulations and may take longer.")
    return errors, warnings


def _geometric_return(values: Iterable[float], fallback: float) -> float:
    clean = np.asarray([float(v) for v in values if np.isfinite(v) and float(v) > -1.0], dtype=float)
    if clean.size == 0:
        return float(fallback)
    return float(np.expm1(np.mean(np.log1p(clean))))


def _safe_correlation(matrix: np.ndarray) -> np.ndarray:
    if matrix.ndim != 2 or matrix.shape[1] == 0:
        return np.eye(1)
    if matrix.shape[0] < 2:
        return np.eye(matrix.shape[1])
    corr = np.corrcoef(matrix, rowvar=False)
    corr = np.atleast_2d(np.asarray(corr, dtype=float))
    if corr.shape != (matrix.shape[1], matrix.shape[1]):
        corr = np.eye(matrix.shape[1])
    corr[~np.isfinite(corr)] = 0.0
    np.fill_diagonal(corr, 1.0)
    return corr


def _positive_semidefinite(matrix: np.ndarray, minimum: float = 1e-10) -> np.ndarray:
    symmetric = (np.asarray(matrix, dtype=float) + np.asarray(matrix, dtype=float).T) / 2.0
    values, vectors = np.linalg.eigh(symmetric)
    values = np.maximum(values, float(minimum))
    return (vectors * values) @ vectors.T


def _monthly_maps(monthly_returns: dict | None) -> dict[str, dict[str, float]]:
    payload = monthly_returns or {}
    maps = payload.get("returns") if isinstance(payload, dict) and "returns" in payload else payload
    if not isinstance(maps, dict):
        return {}
    output: dict[str, dict[str, float]] = {}
    for symbol, values in maps.items():
        if not isinstance(values, dict):
            continue
        clean = {}
        for label, value in values.items():
            parsed = _number(value, float("nan"))
            if np.isfinite(parsed) and parsed > -1.0:
                clean[str(label)] = float(parsed)
        output[str(symbol).upper()] = clean
    return output


def prepare_projection_model(
    market: pd.DataFrame,
    holdings: list[str],
    annual_year_columns: list[str] | tuple[str, ...],
    monthly_returns: dict | None = None,
    use_monthly: bool = False,
    capital_market_return: float | None = None,
) -> PreparedProjectionModel:
    """Build expected returns/covariance and explicit history diagnostics."""

    defaults = model_defaults()
    cma = capital_market_assumptions()["broad_market_annual_geometric_return"]
    broad_return = float(cma["value"] if capital_market_return is None else capital_market_return)
    lookup = _normalized_market(market)
    symbols = [str(s).upper() for s in holdings]
    missing = [s for s in symbols if s not in lookup.index]
    if missing:
        raise ProjectionValidationError("Historical data is unavailable for: " + ", ".join(missing))

    rows: list[pd.Series] = []
    names: list[str] = []
    types: list[str] = []
    categories: list[str] = []
    for symbol in symbols:
        row = lookup.loc[symbol]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[-1]
        rows.append(row)
        names.append(str(row.get("Name") or symbol))
        types.append(str(row.get("Type") or "Stock"))
        categories.append(holding_category(row))

    years = sorted({str(y) for y in annual_year_columns if str(y).isdigit()}, key=int)
    if not years:
        raise ProjectionValidationError("No completed annual-return columns are available for projection calibration.")
    universe = market.copy()
    for year in years:
        universe[year] = pd.to_numeric(universe.get(year), errors="coerce") / 100.0
        universe.loc[(universe[year] <= -1.0) | ~np.isfinite(universe[year]), year] = np.nan
    universe["_symbol"] = universe["Symbol"].astype(str).str.upper()
    if "Type" not in universe.columns:
        universe["Type"] = "Stock"
    universe["_category"] = universe.apply(holding_category, axis=1)

    market_proxy = universe[years].median(axis=0, skipna=True).to_numpy(dtype=float)
    if "SPY" in set(universe["_symbol"]):
        spy = universe.loc[universe["_symbol"].eq("SPY"), years].iloc[-1].to_numpy(dtype=float)
        if int(np.isfinite(spy).sum()) >= defaults["minimum_credible_annual_periods"]:
            market_proxy = np.where(np.isfinite(spy), spy, market_proxy)
    market_proxy = np.where(np.isfinite(market_proxy), market_proxy, broad_return)
    market_geo = _geometric_return(market_proxy, broad_return)
    market_vol = float(np.nanstd(market_proxy, ddof=1)) if len(market_proxy) > 1 else 0.18
    if not np.isfinite(market_vol) or market_vol <= 0:
        market_vol = 0.18

    annual_raw = np.full((len(years), len(symbols)), np.nan, dtype=float)
    annual_filled = np.full_like(annual_raw, np.nan)
    expected_annual = np.zeros(len(symbols), dtype=float)
    annual_vols = np.zeros(len(symbols), dtype=float)
    annual_observed: list[int] = []
    annual_imputed: list[int] = []
    category_geos: list[float] = []
    history_records: list[dict] = []
    warnings: list[str] = []

    for idx, (symbol, row, category) in enumerate(zip(symbols, rows, categories)):
        raw = np.asarray([
            _number(row.get(year), float("nan")) / 100.0 for year in years
        ], dtype=float)
        raw[(raw <= -1.0) | ~np.isfinite(raw)] = np.nan
        annual_raw[:, idx] = raw
        category_rows = universe.loc[universe["_category"].eq(category), years]
        category_proxy = category_rows.median(axis=0, skipna=True).to_numpy(dtype=float) if not category_rows.empty else market_proxy.copy()
        category_proxy = np.where(np.isfinite(category_proxy), category_proxy, market_proxy)
        proxy = 0.60 * category_proxy + 0.40 * market_proxy
        filled = np.where(np.isfinite(raw), raw, proxy)
        filled = np.where(np.isfinite(filled), filled, broad_return)
        annual_filled[:, idx] = filled
        observed = int(np.isfinite(raw).sum())
        annual_observed.append(observed)
        annual_imputed.append(int(len(years) - observed))
        security_geo = _geometric_return(raw, market_geo)
        category_geo = _geometric_return(category_proxy, market_geo)
        category_geos.append(category_geo)
        shrunk_market = (
            defaults["capital_market_anchor_weight"] * broad_return
            + (1.0 - defaults["capital_market_anchor_weight"]) * market_geo
        )
        estimate = (
            shrunk_market
            + defaults["stock_specific_excess_return_weight"] * (security_geo - market_geo)
            + defaults["sector_category_excess_return_weight"] * (category_geo - market_geo)
        )
        expected_annual[idx] = np.clip(
            estimate,
            defaults["expected_annual_geometric_return_floor"],
            defaults["expected_annual_geometric_return_ceiling"],
        )
        security_vol = float(np.nanstd(raw, ddof=1)) if observed >= 2 else float("nan")
        category_flat = category_rows.to_numpy(dtype=float).ravel() if not category_rows.empty else np.asarray([])
        category_vol = float(np.nanstd(category_flat, ddof=1)) if int(np.isfinite(category_flat).sum()) >= 2 else market_vol
        reliability = min(1.0, observed / max(1.0, defaults["minimum_high_confidence_annual_periods"]))
        if not np.isfinite(security_vol) or security_vol <= 0:
            security_vol = category_vol
        annual_vols[idx] = float(np.clip(reliability * security_vol + (1.0 - reliability) * (0.6 * category_vol + 0.4 * market_vol), 0.05, 0.80))
        for year_idx, year in enumerate(years):
            was_observed = bool(np.isfinite(raw[year_idx]))
            history_records.append({
                "Ticker": symbol,
                "Frequency": "Annual",
                "Period": year,
                "Return": float(filled[year_idx]),
                "Return %": float(filled[year_idx]) * 100.0,
                "Data Status": "Observed" if was_observed else "Imputed",
                "Imputation Source": "" if was_observed else "60% sector/category + 40% broad-market proxy",
            })

    annual_corr = _safe_correlation(annual_filled)
    shrink = float(defaults["residual_covariance_shrinkage"])
    shrunk_corr = (1.0 - shrink) * annual_corr + shrink * np.eye(len(symbols))
    annual_cov = _positive_semidefinite(shrunk_corr * np.outer(annual_vols, annual_vols))

    monthly_maps = _monthly_maps(monthly_returns)
    monthly_observed = [0 for _ in symbols]
    monthly_imputed = [0 for _ in symbols]
    estimated_monthly: list[str] = []
    period_frequency = "Monthly" if use_monthly else "Yearly"
    if use_monthly:
        month_labels = sorted({label for symbol in symbols for label in monthly_maps.get(symbol, {})})
        # A minimum calibration frame keeps the fallback explicit and usable even when
        # the durable monthly file has not yet been generated for a recent listing.
        if not month_labels:
            month_labels = [f"Calibration-{i + 1:02d}" for i in range(24)]
        monthly_raw = np.full((len(month_labels), len(symbols)), np.nan, dtype=float)
        for idx, symbol in enumerate(symbols):
            values = monthly_maps.get(symbol, {})
            for month_idx, label in enumerate(month_labels):
                value = values.get(label)
                if value is not None and np.isfinite(value) and float(value) > -1.0:
                    monthly_raw[month_idx, idx] = float(value)
        cross_month = np.nanmedian(monthly_raw, axis=1)
        fallback_month_means = np.expm1(np.log1p(expected_annual) / 12.0)
        monthly_filled = monthly_raw.copy()
        monthly_vols = np.zeros(len(symbols), dtype=float)
        for idx, symbol in enumerate(symbols):
            observed = int(np.isfinite(monthly_raw[:, idx]).sum())
            monthly_observed[idx] = observed
            category_peer_indices = [j for j, category in enumerate(categories) if category == categories[idx] and j != idx]
            for month_idx, label in enumerate(month_labels):
                if np.isfinite(monthly_filled[month_idx, idx]):
                    history_records.append({
                        "Ticker": symbol,
                        "Frequency": "Monthly",
                        "Period": label,
                        "Return": float(monthly_filled[month_idx, idx]),
                        "Return %": float(monthly_filled[month_idx, idx]) * 100.0,
                        "Data Status": "Observed",
                        "Imputation Source": "",
                    })
                    continue
                peer_values = monthly_raw[month_idx, category_peer_indices] if category_peer_indices else np.asarray([])
                peer_values = peer_values[np.isfinite(peer_values)]
                if peer_values.size:
                    proxy_value = float(np.median(peer_values))
                    source = "selected sector/category peer monthly history"
                elif np.isfinite(cross_month[month_idx]):
                    proxy_value = float(cross_month[month_idx])
                    source = "selected-portfolio monthly market proxy"
                else:
                    proxy_value = float(fallback_month_means[idx])
                    source = "calibrated annual volatility and sector/category history"
                monthly_filled[month_idx, idx] = proxy_value
                monthly_imputed[idx] += 1
                history_records.append({
                    "Ticker": symbol,
                    "Frequency": "Monthly",
                    "Period": label,
                    "Return": proxy_value,
                    "Return %": proxy_value * 100.0,
                    "Data Status": "Imputed",
                    "Imputation Source": source,
                })
            observed_vol = float(np.nanstd(monthly_raw[:, idx], ddof=1)) if observed >= 2 else float("nan")
            calibrated_vol = max(0.01, float(annual_vols[idx]) / math.sqrt(12.0))
            reliability = min(1.0, observed / max(1.0, defaults["minimum_high_confidence_monthly_periods"]))
            if not np.isfinite(observed_vol) or observed_vol <= 0:
                observed_vol = calibrated_vol
            monthly_vols[idx] = float(np.clip(reliability * observed_vol + (1.0 - reliability) * calibrated_vol, 0.01, 0.35))
            if observed < len(month_labels) or observed < 12:
                estimated_monthly.append(symbol)
        if estimated_monthly:
            warnings.append(
                "Lower-confidence monthly calibration used for: " + ", ".join(estimated_monthly)
                + ". Estimated periods are identified in Historical Returns Used."
            )
        monthly_corr = _safe_correlation(monthly_filled) if len(month_labels) >= 3 else annual_corr
        shrunk_monthly_corr = (1.0 - shrink) * monthly_corr + shrink * np.eye(len(symbols))
        period_covariance = _positive_semidefinite(shrunk_monthly_corr * np.outer(monthly_vols, monthly_vols))
        period_log_returns = np.log1p(expected_annual) / 12.0
    else:
        period_covariance = annual_cov
        period_log_returns = np.log1p(expected_annual)

    min_annual = min(annual_observed) if annual_observed else 0
    if use_monthly:
        min_monthly = min(monthly_observed) if monthly_observed else 0
        if min_annual >= defaults["minimum_high_confidence_annual_periods"] and min_monthly >= defaults["minimum_high_confidence_monthly_periods"]:
            confidence = "High"
        elif min_annual >= defaults["minimum_credible_annual_periods"] and min_monthly >= 12:
            confidence = "Medium"
        else:
            confidence = "Low"
    elif min_annual >= defaults["minimum_high_confidence_annual_periods"]:
        confidence = "High"
    elif min_annual >= defaults["minimum_credible_annual_periods"]:
        confidence = "Medium"
    else:
        confidence = "Low"
    credible = bool(len(years) >= defaults["minimum_credible_annual_periods"] and sum(annual_observed) >= defaults["minimum_credible_annual_periods"])
    if any(value < defaults["minimum_high_confidence_annual_periods"] for value in annual_observed):
        limited = [f"{symbol} ({count} observed annual periods)" for symbol, count in zip(symbols, annual_observed) if count < defaults["minimum_high_confidence_annual_periods"]]
        warnings.append("Limited-history blend applied to " + ", ".join(limited) + ".")

    holding_records = []
    for idx, symbol in enumerate(symbols):
        holding_records.append({
            "Ticker": symbol,
            "Name": names[idx],
            "Type": types[idx],
            "Sector / ETF Category": categories[idx],
            "Expected Annual Geometric Return": float(expected_annual[idx]),
            "Expected Annual Geometric Return %": float(expected_annual[idx]) * 100.0,
            "Annual Volatility": float(annual_vols[idx]),
            "Annual Volatility %": float(annual_vols[idx]) * 100.0,
            "Observed Annual Periods": int(annual_observed[idx]),
            "Imputed Annual Periods": int(annual_imputed[idx]),
            "Observed Monthly Periods": int(monthly_observed[idx]),
            "Imputed Monthly Periods": int(monthly_imputed[idx]),
            "Monthly Data": "Estimated / blended" if symbol in estimated_monthly else ("Actual" if use_monthly else "Not required"),
        })
    diagnostics = {
        "period_frequency": period_frequency,
        "annual_history_start": years[0],
        "annual_history_end": years[-1],
        "annual_periods_available": len(years),
        "market_proxy_geometric_return": market_geo,
        "market_proxy_volatility": market_vol,
        "covariance_shrinkage": shrink,
        "estimated_monthly_holdings": estimated_monthly,
        "confidence": confidence,
        "credible": credible,
    }
    return PreparedProjectionModel(
        symbols=symbols,
        names=names,
        types=types,
        categories=categories,
        expected_annual_returns=expected_annual,
        annual_covariance=annual_cov,
        period_log_returns=np.asarray(period_log_returns, dtype=float),
        period_covariance=np.asarray(period_covariance, dtype=float),
        period_frequency=period_frequency,
        confidence=confidence,
        credible=credible,
        warnings=warnings,
        holding_assumptions=pd.DataFrame(holding_records),
        historical_returns=pd.DataFrame(history_records),
        diagnostics=diagnostics,
    )


def projection_cache_key(
    inputs: dict,
    market: pd.DataFrame,
    annual_year_columns: list[str] | tuple[str, ...],
    monthly_returns: dict | None = None,
    data_as_of: str = "",
) -> str:
    """Stable digest for identical input/data combinations."""

    normalized = normalize_projection_inputs(inputs)
    holdings = normalized["holdings"]
    lookup = _normalized_market(market)
    data_rows = []
    for symbol in holdings:
        if symbol not in lookup.index:
            continue
        row = lookup.loc[symbol]
        if isinstance(row, pd.DataFrame):
            row = row.iloc[-1]
        data_rows.append({
            "Symbol": symbol,
            "Type": str(row.get("Type") or ""),
            "Category": holding_category(row),
            "Annual": {str(year): _number(row.get(str(year)), float("nan")) for year in annual_year_columns},
        })
    monthly = _monthly_maps(monthly_returns)
    monthly_selected = {symbol: monthly.get(symbol, {}) for symbol in holdings}
    payload = {
        "inputs": normalized,
        "data": data_rows,
        "monthly": monthly_selected,
        "data_as_of": str(data_as_of or ""),
        "model": model_defaults(),
        "capital_market": capital_market_assumptions(),
    }
    encoded = json.dumps(payload, sort_keys=True, default=str, allow_nan=True).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _strategy_names(strategy: str) -> list[str]:
    if strategy == "Rebalanced":
        return ["Rebalanced"]
    if strategy == "Non-Rebalanced":
        return ["Non-Rebalanced"]
    return ["Rebalanced", "Non-Rebalanced"]


def _quantile(values: np.ndarray, percentile: float) -> float:
    values = np.asarray(values, dtype=float)
    finite = values[np.isfinite(values)]
    return float(np.percentile(finite, percentile)) if finite.size else 0.0


def _median(values: np.ndarray) -> float:
    return _quantile(values, 50.0)


def _safe_weights(holdings: np.ndarray) -> np.ndarray:
    totals = holdings.sum(axis=1)
    return np.divide(holdings, totals[:, None], out=np.zeros_like(holdings), where=totals[:, None] > 0)


def _make_state(simulations: int, starting: float, target: np.ndarray) -> dict:
    holdings = np.tile(target * float(starting), (simulations, 1)).astype(float)
    return {
        "holdings": holdings,
        "depleted": np.zeros(simulations, dtype=bool),
        "first_depletion": np.full(simulations, -1, dtype=int),
        "cumulative_withdrawal": np.zeros(simulations),
        "cumulative_contribution": np.zeros(simulations),
        "cumulative_fees": np.zeros(simulations),
        "gross_total": np.zeros(simulations),
        "full_success": np.ones(simulations, dtype=bool),
        "peak": np.full(simulations, float(starting)),
        "max_drawdown": np.zeros(simulations),
        "positive_base": 0,
        "base_observations": 0,
        "positive_year": 0,
        "year_observations": 0,
        "annual_return_medians": [],
        "annual_factor": np.ones(simulations),
        "table_rows": [],
        "detail_rows": [],
        "year_accumulator": None,
    }


def _period_request(inputs: dict, base_index: int, base_monthly: bool) -> float:
    frequency = inputs["withdrawal_frequency"]
    if frequency == "No Withdrawal":
        return 0.0
    year_index = base_index // 12 if base_monthly else base_index
    inflation = inputs["withdrawal_inflation_rate"] if inputs["inflation_adjust_withdrawals"] else 0.0
    inflation_factor = (1.0 + inflation) ** year_index
    if frequency == "Monthly":
        return float(inputs["monthly_withdrawal"]) * inflation_factor
    if not base_monthly:
        return float(inputs["annual_withdrawal"]) * inflation_factor
    month_number = (base_index % 12) + 1
    target_month = 1 if inputs["withdrawal_timing"] == "Beginning of period" else 12
    return float(inputs["annual_withdrawal"]) * inflation_factor if month_number == target_month else 0.0


def _period_contribution(inputs: dict, base_index: int, base_monthly: bool) -> float:
    amount = float(inputs["additional_contribution"])
    if amount <= 0:
        return 0.0
    if inputs["withdrawal_frequency"] == "Monthly":
        return amount
    if not base_monthly:
        return amount
    return amount if (base_index % 12) == 11 else 0.0


def _rebalance_due(frequency: str, base_index: int, base_monthly: bool) -> bool:
    if not base_monthly:
        return True
    month_number = (base_index % 12) + 1
    if frequency == "Monthly":
        return True
    if frequency == "Quarterly":
        return month_number in {3, 6, 9, 12}
    return month_number == 12


def _period_label(start_year: int, base_index: int, base_monthly: bool) -> tuple[str, int, int | None]:
    if base_monthly:
        year = int(start_year + base_index // 12)
        month = int((base_index % 12) + 1)
        return f"{year}-{month:02d}", year, month
    year = int(start_year + base_index)
    return str(year), year, None


def _start_accumulator(state: dict, period_label: str, year: int, month: int | None) -> None:
    h = state["holdings"]
    state["year_accumulator"] = {
        "Period": period_label,
        "Year": year,
        "Month": month,
        "beginning": h.sum(axis=1).copy(),
        "holding_beginning": h.copy(),
        "gross": np.zeros(h.shape[0]),
        "requested": np.zeros(h.shape[0]),
        "actual": np.zeros(h.shape[0]),
        "shortfall": np.zeros(h.shape[0]),
        "contribution": np.zeros(h.shape[0]),
        "fees": np.zeros(h.shape[0]),
        "holding_factor": np.ones_like(h),
        "portfolio_factor": np.ones(h.shape[0]),
        "holding_profit": np.zeros_like(h),
        "holding_withdrawal": np.zeros_like(h),
        "holding_rebalance": np.zeros_like(h),
    }


def _step_state(
    state: dict,
    simulated_returns: np.ndarray,
    request: float,
    contribution: float,
    fee_rate: float,
    timing: str,
    target: np.ndarray,
    rebalanced: bool,
    rebalance_due: bool,
    base_index: int,
) -> dict:
    holdings = state["holdings"]
    simulations = holdings.shape[0]
    beginning_holdings = holdings.copy()
    beginning = beginning_holdings.sum(axis=1)
    active = ~state["depleted"]
    requested = np.full(simulations, max(0.0, request), dtype=float)
    actual = np.zeros(simulations)
    holding_withdrawal = np.zeros_like(holdings)

    if timing == "Beginning of period" and request > 0:
        actual = np.where(active, np.minimum(requested, beginning), 0.0)
        scale = np.divide(beginning - actual, beginning, out=np.zeros_like(beginning), where=beginning > 0)
        post = holdings * scale[:, None]
        holding_withdrawal += holdings - post
        holdings = post
    investable = holdings.copy()
    holdings = np.maximum(0.0, holdings * (1.0 + simulated_returns))
    holding_profit = holdings - investable
    gross_profit = holding_profit.sum(axis=1)

    if timing == "End of period" and request > 0:
        available = holdings.sum(axis=1)
        actual = np.where(active, np.minimum(requested, available), 0.0)
        scale = np.divide(available - actual, available, out=np.zeros_like(available), where=available > 0)
        post = holdings * scale[:, None]
        holding_withdrawal += holdings - post
        holdings = post

    shortfall = np.maximum(0.0, requested - actual)
    state["full_success"] &= shortfall <= 0.005
    after_withdrawal = holdings.sum(axis=1)
    newly_depleted = active & (request > 0) & (after_withdrawal <= 0.005)
    state["depleted"] |= newly_depleted
    state["first_depletion"] = np.where(
        newly_depleted & (state["first_depletion"] < 0), base_index, state["first_depletion"]
    )

    contribution_values = np.where(~state["depleted"], max(0.0, contribution), 0.0)
    if contribution > 0:
        holdings += contribution_values[:, None] * target[None, :]

    before_fee = holdings.sum(axis=1)
    fees = np.where(~state["depleted"], np.minimum(before_fee, before_fee * max(0.0, fee_rate)), 0.0)
    fee_scale = np.divide(before_fee - fees, before_fee, out=np.zeros_like(before_fee), where=before_fee > 0)
    holdings *= fee_scale[:, None]

    rebalancing_trade = np.zeros_like(holdings)
    if rebalanced and rebalance_due:
        totals = holdings.sum(axis=1)
        target_holdings = totals[:, None] * target[None, :]
        rebalancing_trade = target_holdings - holdings
        holdings = target_holdings
    holdings[state["depleted"], :] = 0.0
    ending = holdings.sum(axis=1)

    state["holdings"] = holdings
    state["cumulative_withdrawal"] += actual
    state["cumulative_contribution"] += contribution_values
    state["cumulative_fees"] += fees
    state["gross_total"] += gross_profit
    state["peak"] = np.maximum(state["peak"], ending)
    drawdown = np.divide(ending, state["peak"], out=np.zeros_like(ending), where=state["peak"] > 0) - 1.0
    state["max_drawdown"] = np.minimum(state["max_drawdown"], drawdown)
    portfolio_return = np.divide(
        gross_profit,
        investable.sum(axis=1),
        out=np.zeros_like(gross_profit),
        where=investable.sum(axis=1) > 0,
    )
    state["positive_base"] += int(np.sum(portfolio_return > 0))
    state["base_observations"] += simulations
    state["annual_factor"] *= 1.0 + portfolio_return
    return {
        "beginning": beginning,
        "gross": gross_profit,
        "requested": requested,
        "actual": actual,
        "shortfall": shortfall,
        "contribution": contribution_values,
        "fees": fees,
        "ending": ending,
        "net_change": ending - beginning,
        "portfolio_return": portfolio_return,
        "holding_beginning": beginning_holdings,
        "holding_return": simulated_returns,
        "holding_profit": holding_profit,
        "holding_withdrawal": holding_withdrawal,
        "holding_ending": holdings.copy(),
        "holding_beginning_weight": _safe_weights(beginning_holdings),
        "holding_ending_weight": _safe_weights(holdings),
        "holding_rebalance": rebalancing_trade,
    }


def _accumulate(state: dict, metrics: dict) -> None:
    acc = state["year_accumulator"]
    acc["gross"] += metrics["gross"]
    acc["requested"] += metrics["requested"]
    acc["actual"] += metrics["actual"]
    acc["shortfall"] += metrics["shortfall"]
    acc["contribution"] += metrics["contribution"]
    acc["fees"] += metrics["fees"]
    acc["holding_factor"] *= 1.0 + metrics["holding_return"]
    acc["portfolio_factor"] *= 1.0 + metrics["portfolio_return"]
    acc["holding_profit"] += metrics["holding_profit"]
    acc["holding_withdrawal"] += metrics["holding_withdrawal"]
    acc["holding_rebalance"] += metrics["holding_rebalance"]


def _finish_output_period(
    state: dict,
    symbols: list[str],
    starting_investment: float,
    show_month: bool,
) -> None:
    acc = state["year_accumulator"]
    ending = state["holdings"].sum(axis=1)
    net_change = ending - acc["beginning"]
    period_return = acc["holding_factor"] - 1.0
    portfolio_return = acc["portfolio_factor"] - 1.0
    wealth_profit = (
        ending + state["cumulative_withdrawal"] - state["cumulative_contribution"] - float(starting_investment)
    )
    cumulative_wealth = ending + state["cumulative_withdrawal"]
    row = {
        "Period": acc["Period"],
        "Year": int(acc["Year"]),
        "Beginning Balance": _median(acc["beginning"]),
        "P10 Gross Profit": _quantile(acc["gross"], 10),
        "Median Gross Profit": _median(acc["gross"]),
        "P90 Gross Profit": _quantile(acc["gross"], 90),
        "Median Portfolio Return": _median(portfolio_return) * 100.0,
        "Requested Withdrawal": _median(acc["requested"]),
        "Actual Withdrawal": _median(acc["actual"]),
        "Withdrawal Shortfall": _median(acc["shortfall"]),
        "Additional Contribution": _median(acc["contribution"]),
        "Fees": _median(acc["fees"]),
        "Median Net Change": _median(net_change),
        "P5 Ending Balance": _quantile(ending, 5),
        "P10 Ending Balance": _quantile(ending, 10),
        "Median Ending Balance": _median(ending),
        "P90 Ending Balance": _quantile(ending, 90),
        "P95 Ending Balance": _quantile(ending, 95),
        "Cumulative Withdrawals": _median(state["cumulative_withdrawal"]),
        "P5 Cumulative Wealth": _quantile(cumulative_wealth, 5),
        "P10 Cumulative Wealth": _quantile(cumulative_wealth, 10),
        "Cumulative Wealth": _median(cumulative_wealth),
        "P90 Cumulative Wealth": _quantile(cumulative_wealth, 90),
        "P95 Cumulative Wealth": _quantile(cumulative_wealth, 95),
        "Total Wealth Profit": _median(wealth_profit),
        "Depletion Probability": float(np.mean(state["depleted"]) * 100.0),
    }
    if show_month:
        row["Month"] = int(acc["Month"])
        row["Median Monthly Return"] = row.pop("Median Portfolio Return")
        row["Contribution"] = row.pop("Additional Contribution")
    if row["Median Ending Balance"] <= 0.005:
        row["Status"] = "Depleted"
    elif row["Withdrawal Shortfall"] > 0.005:
        row["Status"] = "Partial Withdrawal"
    else:
        row["Status"] = "Active"
    state["table_rows"].append(row)

    beginning_weights = _safe_weights(acc["holding_beginning"])
    ending_weights = _safe_weights(state["holdings"])
    for idx, symbol in enumerate(symbols):
        detail = {
            "Period": acc["Period"],
            "Year": int(acc["Year"]),
            "Ticker": symbol,
            "Beginning Value": _median(acc["holding_beginning"][:, idx]),
            "Period Return": _median(period_return[:, idx]) * 100.0,
            "Profit or Loss": _median(acc["holding_profit"][:, idx]),
            "Withdrawal Funded by Holding": _median(acc["holding_withdrawal"][:, idx]),
            "Ending Value": _median(state["holdings"][:, idx]),
            "Beginning Weight": _median(beginning_weights[:, idx]) * 100.0,
            "Ending Weight": _median(ending_weights[:, idx]) * 100.0,
            "Rebalancing Buy or Sell Amount": _median(acc["holding_rebalance"][:, idx]),
        }
        if show_month:
            detail["Month"] = int(acc["Month"])
        state["detail_rows"].append(detail)
    state["year_accumulator"] = None


def _complete_year(state: dict) -> None:
    annual_return = state["annual_factor"] - 1.0
    state["annual_return_medians"].append(_median(annual_return))
    state["positive_year"] += int(np.sum(annual_return > 0))
    state["year_observations"] += len(annual_return)
    state["annual_factor"] = np.ones_like(state["annual_factor"])


def _depletion_label(first_indices: np.ndarray, start_year: int, base_monthly: bool) -> tuple[str | None, int | None, str | None]:
    depleted = first_indices[first_indices >= 0]
    if depleted.size == 0:
        return None, None, None
    median_index = int(np.rint(np.median(depleted)))
    label, year, month = _period_label(start_year, median_index, base_monthly)
    month_year = label if month is not None else None
    return label, year, month_year


def _finalize_strategy(
    state: dict,
    no_withdrawal_state: dict | None,
    inputs: dict,
    start_year: int,
    base_monthly: bool,
    model: PreparedProjectionModel,
) -> dict:
    ending = state["holdings"].sum(axis=1)
    total_wealth_profit = ending + state["cumulative_withdrawal"] - state["cumulative_contribution"] - inputs["starting_investment"]
    depletion_label, depletion_year, depletion_month_year = _depletion_label(
        state["first_depletion"], start_year, base_monthly
    )
    annual_medians = list(state["annual_return_medians"])
    best_index = int(np.argmax(annual_medians)) if annual_medians else 0
    worst_index = int(np.argmin(annual_medians)) if annual_medians else 0
    no_withdrawal_ending = np.zeros(1)
    if no_withdrawal_state is not None:
        no_withdrawal_ending = no_withdrawal_state["holdings"].sum(axis=1)
    years = max(1, int(inputs["future_years"]))
    cagr = np.where(
        no_withdrawal_ending > 0,
        np.power(no_withdrawal_ending / float(inputs["starting_investment"]), 1.0 / years) - 1.0,
        -1.0,
    )
    summary = {
        "Starting Investment": float(inputs["starting_investment"]),
        "Forecast Period": f"{start_year}-{start_year + years - 1}",
        "Selected Holdings": ", ".join(inputs["holdings"]),
        "Allocation": ", ".join(f"{s} {inputs['allocations'][s]:.2f}%" for s in inputs["holdings"]),
        "Median Ending Balance": _median(ending),
        "P10 Ending Balance": _quantile(ending, 10),
        "P90 Ending Balance": _quantile(ending, 90),
        "P5 Ending Balance": _quantile(ending, 5),
        "P95 Ending Balance": _quantile(ending, 95),
        "Median Total Investment Profit": _median(state["gross_total"]),
        "Median Total Wealth Profit": _median(total_wealth_profit),
        "Median Actual Withdrawals Received": _median(state["cumulative_withdrawal"]),
        "Withdrawal Shortfall": max(
            0.0,
            sum(float(row["Requested Withdrawal"]) for row in state["table_rows"])
            - _median(state["cumulative_withdrawal"]),
        ),
        "Full-Withdrawal Success Probability": float(np.mean(state["full_success"]) * 100.0),
        "Depletion Probability": float(np.mean(state["first_depletion"] >= 0) * 100.0),
        "Median Depletion Year": depletion_year,
        "Median Depletion Month and Year": depletion_month_year,
        "Median Depletion Period": depletion_label,
        "Median No-Withdrawal Ending Balance": _median(no_withdrawal_ending) if no_withdrawal_state is not None else None,
        "Median No-Withdrawal CAGR": _median(cagr) * 100.0 if no_withdrawal_state is not None else None,
        "Best Modeled Year": f"{start_year + best_index} ({annual_medians[best_index] * 100.0:+.2f}%)" if annual_medians else "N/A",
        "Worst Modeled Year": f"{start_year + worst_index} ({annual_medians[worst_index] * 100.0:+.2f}%)" if annual_medians else "N/A",
        "Maximum Projected Drawdown": _median(state["max_drawdown"]) * 100.0,
        "Positive-Year Percentage": (100.0 * state["positive_year"] / state["year_observations"]) if state["year_observations"] else 0.0,
        "Positive-Month Percentage": (100.0 * state["positive_base"] / state["base_observations"]) if base_monthly and state["base_observations"] else None,
        "Model Confidence": model.confidence,
    }
    return {
        "summary": summary,
        "table": pd.DataFrame(state["table_rows"]),
        "holding_details": pd.DataFrame(state["detail_rows"]),
        "chart": pd.DataFrame(state["table_rows"]),
        "no_withdrawal_chart": (
            pd.DataFrame(no_withdrawal_state["table_rows"])
            if no_withdrawal_state is not None else pd.DataFrame()
        ),
    }


def run_future_projection(
    market: pd.DataFrame,
    inputs: dict,
    annual_year_columns: list[str] | tuple[str, ...],
    monthly_returns: dict | None = None,
    data_as_of: str = "",
    model_as_of: str | None = None,
    progress: ProgressCallback | None = None,
) -> dict:
    """Run the complete Future Projection with one shared return cube per strategy.

    Rebalanced and Non-Rebalanced paths consume identical random returns, making
    their side-by-side difference attributable to portfolio maintenance rather
    than different random draws.
    """

    normalized = normalize_projection_inputs(inputs)
    errors, input_warnings = validate_projection_inputs(normalized, market)
    if errors:
        raise ProjectionValidationError(" ".join(errors))
    base_monthly = (
        normalized["withdrawal_frequency"] == "Monthly"
        or (
            normalized["strategy"] in {"Rebalanced", "Both"}
            and normalized["rebalancing_frequency"] in {"Quarterly", "Monthly"}
        )
    )
    model = prepare_projection_model(
        market,
        normalized["holdings"],
        annual_year_columns,
        monthly_returns=monthly_returns,
        use_monthly=base_monthly,
    )
    if not model.credible:
        raise ProjectionValidationError(
            "The selected holdings have too little completed historical data for a credible projection. "
            "Choose holdings with at least three observed completed annual periods."
        )

    defaults = model_defaults()
    simulations = int(normalized["simulation_count"])
    start_year = int(normalized["forecast_start_year"])
    base_periods = int(normalized["future_years"] * (12 if base_monthly else 1))
    output_monthly = normalized["withdrawal_frequency"] == "Monthly"
    target = np.asarray([normalized["allocations"][s] / 100.0 for s in normalized["holdings"]], dtype=float)
    target = target / target.sum()
    rng = np.random.default_rng(int(normalized["random_seed"]))
    covariance = _positive_semidefinite(model.period_covariance)
    chol = np.linalg.cholesky(covariance)
    df = float(defaults["student_t_degrees_of_freedom"])
    t_variance_scale = math.sqrt((df - 2.0) / df)
    transition = np.asarray(defaults["regime_transition_matrix"], dtype=float)
    transition_cumulative = np.cumsum(transition, axis=1)
    regime = rng.choice(3, size=simulations, p=np.asarray(defaults["initial_regime_probabilities"], dtype=float))
    regime_names = defaults["regime_order"]
    annual_adjustments = np.asarray([defaults["regime_annual_log_return_adjustments"][name] for name in regime_names])
    vol_multipliers = np.asarray([defaults["regime_volatility_multipliers"][name] for name in regime_names])
    state_names = _strategy_names(normalized["strategy"])
    states = {
        name: _make_state(simulations, normalized["starting_investment"], target)
        for name in state_names
    }
    no_withdrawal_states = {
        name: _make_state(simulations, normalized["starting_investment"], target)
        for name in state_names
    } if normalized["include_no_withdrawal_comparison"] else {}
    regime_counts = np.zeros(3, dtype=np.int64)
    annual_fee = float(normalized["annual_management_fee"])
    fee_rate = math.expm1(math.log1p(annual_fee) / 12.0) if base_monthly else annual_fee

    for base_index in range(base_periods):
        if base_index:
            u = rng.random(simulations)
            next_regime = regime.copy()
            for current in range(3):
                mask = regime == current
                next_regime[mask] = np.searchsorted(transition_cumulative[current], u[mask], side="right")
            regime = next_regime
        regime_counts += np.bincount(regime, minlength=3)
        shocks = rng.standard_t(df, size=(simulations, len(target))) * t_variance_scale
        correlated = shocks @ chol.T
        adjustment = annual_adjustments[regime] / (12.0 if base_monthly else 1.0)
        simulated_log = (
            model.period_log_returns[None, :]
            + adjustment[:, None]
            + correlated * vol_multipliers[regime, None]
        )
        simulated_returns = np.expm1(simulated_log)
        lower = defaults["individual_monthly_return_floor"] if base_monthly else defaults["individual_annual_return_floor"]
        upper = defaults["individual_monthly_return_ceiling"] if base_monthly else defaults["individual_annual_return_ceiling"]
        simulated_returns = np.clip(simulated_returns, lower, upper)
        request = _period_request(normalized, base_index, base_monthly)
        contribution = _period_contribution(normalized, base_index, base_monthly)
        period_label, year, month = _period_label(start_year, base_index, base_monthly)
        output_boundary = (output_monthly or not base_monthly or (base_index % 12) == 11)
        output_start = output_monthly or not base_monthly or (base_index % 12) == 0

        for name, state in states.items():
            if output_start:
                display_label = period_label if output_monthly else str(year)
                _start_accumulator(state, display_label, year, month if output_monthly else None)
            metrics = _step_state(
                state,
                simulated_returns,
                request,
                contribution,
                fee_rate,
                normalized["withdrawal_timing"],
                target,
                name == "Rebalanced",
                _rebalance_due(normalized["rebalancing_frequency"], base_index, base_monthly),
                base_index,
            )
            _accumulate(state, metrics)
            if output_boundary:
                _finish_output_period(state, model.symbols, normalized["starting_investment"], output_monthly)

        for name, state in no_withdrawal_states.items():
            if output_start:
                display_label = period_label if output_monthly else str(year)
                _start_accumulator(state, display_label, year, month if output_monthly else None)
            metrics = _step_state(
                state,
                simulated_returns,
                0.0,
                contribution,
                fee_rate,
                "End of period",
                target,
                name == "Rebalanced",
                _rebalance_due(normalized["rebalancing_frequency"], base_index, base_monthly),
                base_index,
            )
            _accumulate(state, metrics)
            if output_boundary:
                _finish_output_period(state, model.symbols, normalized["starting_investment"], output_monthly)
        if (not base_monthly) or (base_index % 12) == 11:
            for state in states.values():
                _complete_year(state)
            for state in no_withdrawal_states.values():
                _complete_year(state)
        if progress:
            estimated_completed = int(round(simulations * (base_index + 1) / base_periods))
            progress(estimated_completed, simulations, f"Modeling period {base_index + 1:,} of {base_periods:,}")

    strategies = {
        name: _finalize_strategy(
            state,
            no_withdrawal_states.get(name),
            normalized,
            start_year,
            base_monthly,
            model,
        )
        for name, state in states.items()
    }
    comparison = pd.DataFrame()
    if {"Rebalanced", "Non-Rebalanced"}.issubset(strategies):
        rb = strategies["Rebalanced"]["table"]
        nr = strategies["Non-Rebalanced"]["table"]
        comparison = pd.DataFrame({
            "Period": rb["Period"],
            "Rebalanced Median Ending Balance": rb["Median Ending Balance"],
            "Non-Rebalanced Median Ending Balance": nr["Median Ending Balance"],
            "Difference": rb["Median Ending Balance"] - nr["Median Ending Balance"],
            "Rebalanced Depletion Probability": rb["Depletion Probability"],
            "Non-Rebalanced Depletion Probability": nr["Depletion Probability"],
        })

    assumption_rows = []
    for key, entry in capital_market_assumptions().items():
        assumption_rows.append({
            "Assumption": key.replace("_", " ").title(),
            "Value": entry["value"],
            "Source": entry["source"],
            "As-of Date": entry["as_of_date"],
            "Last Updated Date": entry["last_updated_date"],
            "Description": entry["description"],
        })
    assumption_rows.extend([
        {"Assumption": "Student-t degrees of freedom", "Value": df, "Source": "MarketScope model configuration", "As-of Date": normalized["capital_market_assumption_date"], "Last Updated Date": model_as_of or date.today().isoformat(), "Description": "Heavy-tailed correlated shocks."},
        {"Assumption": "Stock-specific excess-return weight", "Value": defaults["stock_specific_excess_return_weight"], "Source": "MarketScope model configuration", "As-of Date": normalized["capital_market_assumption_date"], "Last Updated Date": model_as_of or date.today().isoformat(), "Description": "Weight on a holding's historical excess-return signal."},
        {"Assumption": "Sector/category excess-return weight", "Value": defaults["sector_category_excess_return_weight"], "Source": "MarketScope model configuration", "As-of Date": normalized["capital_market_assumption_date"], "Last Updated Date": model_as_of or date.today().isoformat(), "Description": "Weight on sector or ETF-category excess return."},
        {"Assumption": "Residual covariance shrinkage", "Value": defaults["residual_covariance_shrinkage"], "Source": "MarketScope model configuration", "As-of Date": normalized["capital_market_assumption_date"], "Last Updated Date": model_as_of or date.today().isoformat(), "Description": "Sample correlation shrunk toward independent residuals."},
        {"Assumption": "Regimes", "Value": "Bear / Normal / Bull", "Source": "MarketScope model configuration", "As-of Date": normalized["capital_market_assumption_date"], "Last Updated Date": model_as_of or date.today().isoformat(), "Description": "Markov-style transitions with regime-dependent drift and volatility."},
    ])
    diagnostics = dict(model.diagnostics)
    diagnostics.update({
        "simulation_count": simulations,
        "random_seed": normalized["random_seed"],
        "base_frequency": "Monthly" if base_monthly else "Yearly",
        "output_frequency": "Monthly" if output_monthly else "Yearly",
        "regime_observation_counts": {name: int(regime_counts[idx]) for idx, name in enumerate(regime_names)},
        "deterministic": True,
        "chart_table_shared_source": True,
    })
    warnings = list(dict.fromkeys([*input_warnings, *model.warnings]))
    return {
        "inputs": normalized,
        "metadata": {
            "data_as_of": str(data_as_of or "Not available"),
            "model_as_of": str(model_as_of or date.today().isoformat()),
            "forecast_start_year": start_year,
            "forecast_end_year": start_year + int(normalized["future_years"]) - 1,
            "simulation_count": simulations,
            "random_seed": normalized["random_seed"],
            "base_frequency": "Monthly" if base_monthly else "Yearly",
            "output_frequency": "Monthly" if output_monthly else "Yearly",
        },
        "strategies": strategies,
        "comparison": comparison,
        "model_assumptions": pd.DataFrame(assumption_rows),
        "holding_assumptions": model.holding_assumptions,
        "historical_returns": model.historical_returns,
        "diagnostics": diagnostics,
        "warnings": warnings,
        "limitations": list(MODEL_LIMITATIONS),
        "cache_key": projection_cache_key(normalized, market, annual_year_columns, monthly_returns, data_as_of),
    }


def _excel_safe_frame(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    for column in out.columns:
        if out[column].dtype == "object":
            out[column] = out[column].map(lambda value: json.dumps(value, sort_keys=True) if isinstance(value, (dict, list)) else value)
    return out


def build_csv_export(result: dict) -> bytes:
    """Create one portable long-form CSV from the exact table/chart source."""

    frames = []
    for strategy, payload in (result.get("strategies") or {}).items():
        frame = payload.get("table", pd.DataFrame()).copy()
        if not frame.empty:
            frame.insert(0, "Strategy", strategy)
            frames.append(frame)
    combined = pd.concat(frames, ignore_index=True, sort=False) if frames else pd.DataFrame()
    return combined.to_csv(index=False, float_format="%.10f").encode("utf-8")


def build_excel_export(result: dict) -> bytes:
    """Create the required multi-sheet, source-backed Excel workbook."""

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    inputs = result.get("inputs") or {}
    summary_rows = []
    for strategy, payload in (result.get("strategies") or {}).items():
        for metric, value in (payload.get("summary") or {}).items():
            summary_rows.append({"Strategy": strategy, "Metric": metric, "Value": value})
    holdings = []
    for symbol in inputs.get("holdings", []):
        holdings.append({
            "Ticker": symbol,
            "Allocation %": (inputs.get("allocations") or {}).get(symbol),
        })
    diagnostics = pd.DataFrame([
        {"Diagnostic": key, "Value": json.dumps(value, sort_keys=True) if isinstance(value, (dict, list)) else value}
        for key, value in (result.get("diagnostics") or {}).items()
    ])
    sources = pd.DataFrame({
        "Type": ["Warning"] * len(result.get("warnings") or []) + ["Limitation"] * len(result.get("limitations") or []),
        "Text": list(result.get("warnings") or []) + list(result.get("limitations") or []),
    })
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Executive Summary", index=False)
        _excel_safe_frame(result.get("model_assumptions", pd.DataFrame())).to_excel(writer, sheet_name="Model Assumptions", index=False)
        pd.DataFrame(holdings).merge(
            result.get("holding_assumptions", pd.DataFrame()), on="Ticker", how="left"
        ).to_excel(writer, sheet_name="Selected Holdings", index=False)
        strategies = result.get("strategies") or {}
        _excel_safe_frame(strategies.get("Rebalanced", {}).get("table", pd.DataFrame())).to_excel(
            writer, sheet_name="Rebalanced Projection", index=False
        )
        _excel_safe_frame(strategies.get("Non-Rebalanced", {}).get("table", pd.DataFrame())).to_excel(
            writer, sheet_name="Non-Rebalanced Projection", index=False
        )
        _excel_safe_frame(result.get("comparison", pd.DataFrame())).to_excel(writer, sheet_name="Side-by-Side", index=False)
        _excel_safe_frame(result.get("holding_assumptions", pd.DataFrame())).to_excel(writer, sheet_name="Stock ETF Assumptions", index=False)
        _excel_safe_frame(result.get("historical_returns", pd.DataFrame())).to_excel(writer, sheet_name="Historical Returns", index=False)
        diagnostics.to_excel(writer, sheet_name="Model Diagnostics", index=False)
        sources.to_excel(writer, sheet_name="Sources Limitations", index=False)
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="123B68")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            for idx, column_cells in enumerate(sheet.columns, start=1):
                values = [str(cell.value or "") for cell in list(column_cells)[:200]]
                width = min(42, max(11, max((len(value) for value in values), default=10) + 2))
                sheet.column_dimensions[get_column_letter(idx)].width = width
    return output.getvalue()


def build_pdf_export(result: dict, title: str = "MarketScope Future Projection") -> bytes:
    """Create a vector PDF executive report with the required risk disclosure."""

    from reportlab.graphics.charts.linecharts import HorizontalLineChart
    from reportlab.graphics.shapes import Drawing, String
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_LEFT
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.platypus import KeepTogether, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=letter,
        rightMargin=0.42 * inch,
        leftMargin=0.42 * inch,
        topMargin=0.42 * inch,
        bottomMargin=0.42 * inch,
        title=title,
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="MSBody", parent=styles["BodyText"], fontSize=8.3, leading=10.5, textColor=colors.HexColor("#1B2A3A"), alignment=TA_LEFT))
    styles.add(ParagraphStyle(name="MSWarning", parent=styles["BodyText"], fontSize=8, leading=10, textColor=colors.HexColor("#7C2D12"), backColor=colors.HexColor("#FFF7ED"), borderPadding=6))
    styles.add(ParagraphStyle(name="MSCell", parent=styles["BodyText"], fontSize=6.2, leading=7.4, textColor=colors.HexColor("#1B2A3A")))
    story = [Paragraph(title, styles["Title"])]
    metadata = result.get("metadata") or {}
    story.extend([
        Paragraph(
            f"Data as of: {metadata.get('data_as_of', 'N/A')} &nbsp;&nbsp; Model as of: {metadata.get('model_as_of', 'N/A')} &nbsp;&nbsp; "
            f"Simulations: {int(metadata.get('simulation_count') or 0):,} &nbsp;&nbsp; Seed: {metadata.get('random_seed', 'N/A')}",
            styles["MSBody"],
        ),
        Spacer(1, 6),
        Paragraph(
            "Future projections are hypothetical estimates based on historical data and model assumptions. "
            "They are not guaranteed results or individualized investment advice.",
            styles["MSWarning"],
        ),
        Spacer(1, 8),
    ])
    inputs = result.get("inputs") or {}
    input_rows = [
        ["Input", "Value"],
        ["Starting investment", f"${float(inputs.get('starting_investment') or 0):,.2f}"],
        ["Forecast", f"{metadata.get('forecast_start_year')} - {metadata.get('forecast_end_year')}"],
        ["Withdrawal", f"{inputs.get('withdrawal_frequency')} / timing: {inputs.get('withdrawal_timing')}"],
        ["Annual withdrawal", f"${float(inputs.get('annual_withdrawal') or 0):,.2f}"],
        ["Monthly withdrawal", f"${float(inputs.get('monthly_withdrawal') or 0):,.2f}"],
        ["Strategy", str(inputs.get("strategy"))],
        ["Rebalancing", str(inputs.get("rebalancing_frequency"))],
        ["Management fee", f"{float(inputs.get('annual_management_fee') or 0) * 100:.3f}%"],
    ]
    holdings_rows = [["Ticker", "Allocation"]] + [
        [symbol, f"{float((inputs.get('allocations') or {}).get(symbol) or 0):.2f}%"]
        for symbol in inputs.get("holdings", [])
    ]
    left = Table(input_rows, colWidths=[1.55 * inch, 2.55 * inch])
    right = Table(holdings_rows, colWidths=[1.25 * inch, 1.25 * inch])
    for table in (left, right):
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123B68")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
    story.extend([Table([[left, right]], colWidths=[4.25 * inch, 2.65 * inch]), Spacer(1, 10)])

    for strategy, payload in (result.get("strategies") or {}).items():
        summary = payload.get("summary") or {}
        summary_metrics = [
            "Median Ending Balance", "P10 Ending Balance", "P90 Ending Balance",
            "Median Total Wealth Profit", "Median Actual Withdrawals Received", "Withdrawal Shortfall",
            "Full-Withdrawal Success Probability", "Depletion Probability", "Median Depletion Year",
            "Median Depletion Month and Year", "Maximum Projected Drawdown", "Model Confidence",
        ]
        rows = [["Metric", "Result"]]
        for metric in summary_metrics:
            value = summary.get(metric)
            if isinstance(value, (int, float, np.number)) and value is not None:
                if "Probability" in metric or "Drawdown" in metric:
                    formatted = f"{float(value):,.2f}%"
                elif "Year" == metric[-4:]:
                    formatted = str(value)
                else:
                    formatted = f"${float(value):,.2f}"
            else:
                formatted = "N/A" if value in {None, ""} else str(value)
            rows.append([metric, formatted])
        table = Table(rows, colWidths=[3.15 * inch, 3.65 * inch], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123B68")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ]))
        story.append(KeepTogether([
            Paragraph(f"{strategy} summary", styles["Heading2"]),
            table,
            Spacer(1, 8),
        ]))

    strategies = result.get("strategies") or {}
    chart_series = []
    chart_labels = []
    for name in ("Rebalanced", "Non-Rebalanced"):
        frame = strategies.get(name, {}).get("table", pd.DataFrame())
        if not frame.empty:
            chart_series.append(frame["Median Ending Balance"].astype(float).tolist())
            if not chart_labels:
                chart_labels = frame["Period"].astype(str).tolist()
    if chart_series:
        drawing = Drawing(500, 190)
        chart = HorizontalLineChart()
        chart.x = 48
        chart.y = 30
        chart.height = 135
        chart.width = 420
        chart.data = chart_series
        chart.lines[0].strokeColor = colors.HexColor("#2F80ED")
        if len(chart_series) > 1:
            chart.lines[1].strokeColor = colors.HexColor("#F59E0B")
        chart.categoryAxis.categoryNames = chart_labels
        chart.categoryAxis.labels.fontSize = 6
        chart.categoryAxis.labels.angle = 30 if len(chart_labels) > 10 else 0
        chart.valueAxis.valueMin = 0
        chart.valueAxis.labels.fontSize = 6
        drawing.add(chart)
        drawing.add(String(48, 174, "Median projected portfolio balance", fontSize=9, fillColor=colors.HexColor("#123B68")))
        drawing.add(String(295, 174, "Blue: Rebalanced", fontSize=7, fillColor=colors.HexColor("#2F80ED")))
        if len(chart_series) > 1:
            drawing.add(String(385, 174, "Orange: Non-Rebalanced", fontSize=7, fillColor=colors.HexColor("#F59E0B")))
        story.extend([Paragraph("Performance chart", styles["Heading2"]), drawing, Spacer(1, 8)])

    comparison = result.get("comparison", pd.DataFrame())
    if isinstance(comparison, pd.DataFrame) and not comparison.empty:
        compact = comparison.copy()
        if len(compact) > 15:
            indices = np.unique(np.linspace(0, len(compact) - 1, 15).round().astype(int))
            compact = compact.iloc[indices]
        rows = [["Period", "RB ending", "NR ending", "Difference", "RB depletion", "NR depletion"]]
        for _, row in compact.iterrows():
            rows.append([
                str(row["Period"]),
                f"${float(row['Rebalanced Median Ending Balance']):,.0f}",
                f"${float(row['Non-Rebalanced Median Ending Balance']):,.0f}",
                f"${float(row['Difference']):+,.0f}",
                f"{float(row['Rebalanced Depletion Probability']):.1f}%",
                f"{float(row['Non-Rebalanced Depletion Probability']):.1f}%",
            ])
        table = Table(rows, repeatRows=1, colWidths=[0.8 * inch, 1.15 * inch, 1.15 * inch, 1.1 * inch, 1.05 * inch, 1.05 * inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123B68")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
        ]))
        story.append(KeepTogether([
            Paragraph("Rebalanced and non-rebalanced comparison", styles["Heading2"]),
            table,
            Spacer(1, 8),
        ]))

    story.append(PageBreak())
    story.append(Paragraph("Major risk warnings, sources, and limitations", styles["Heading2"]))
    for warning in result.get("warnings") or []:
        story.append(Paragraph("Warning: " + str(warning), styles["MSWarning"]))
        story.append(Spacer(1, 4))
    for limitation in result.get("limitations") or []:
        story.append(Paragraph("- " + str(limitation), styles["MSBody"]))
    story.extend([Spacer(1, 8), Paragraph("Model assumptions", styles["Heading2"])])
    assumptions = result.get("model_assumptions", pd.DataFrame())
    if isinstance(assumptions, pd.DataFrame) and not assumptions.empty:
        rows = [["Assumption", "Value", "Source", "As-of"]]
        for _, row in assumptions.iterrows():
            value = row.get("Value")
            if isinstance(value, float):
                value = f"{value:.4f}"
            rows.append([
                Paragraph(str(row.get("Assumption")), styles["MSCell"]),
                Paragraph(str(value), styles["MSCell"]),
                Paragraph(str(row.get("Source")), styles["MSCell"]),
                Paragraph(str(row.get("As-of Date")), styles["MSCell"]),
            ])
        table = Table(rows, repeatRows=1, colWidths=[1.85 * inch, 0.85 * inch, 3.15 * inch, 0.95 * inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#123B68")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 6.5),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(table)
    doc.build(story)
    return output.getvalue()
